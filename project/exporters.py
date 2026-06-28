# -*- coding: utf-8 -*-
"""
Excel Reporting and Export Engine.
Generates highly styled, professional multi-sheet Excel files with column auto-widths,
freeze panes, headers, and colored highlights.
"""

import pandas as pd
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from .utils import logger

class ExcelExporter:
    COLUMNS = [
        "Job Name", "JD", "Link of the Job", "Area", "Link to Apply", 
        "Place You Got That Job From", "Salary Quotation if Given",
        "Company Name", "Job ID", "Work Mode", "Posted Date", "Freshness Hours", 
        "Experience Required", "Tech Stack", "Skills Required", "Employment Type", 
        "Seniority Level", "Responsibilities", "Benefits", "Recruiter Name", 
        "Recruiter Email", "Hiring Manager Name", "Recruiter LinkedIn URL", 
        "Cold Email Subject", "Custom Cold Email", "Recruiter Contact Strategy"
    ]

    @classmethod
    def export_to_excel(cls, jobs, filename="jobs.xlsx"):
        """
        Exports job data to a styled Excel file with city-specific sheets.
        """
        logger.info(f"Exporter: Writing {len(jobs)} jobs to Excel workbook: {filename}")
        
        # Convert dictionaries to DataFrame
        # Map our internal keys to requested Excel column names
        data_rows = []
        for job in jobs:
            direct_link = job.get("url", "")
            data_rows.append({
                "Job Name": job.get("title", ""),
                "JD": job.get("summary") or job.get("description") or "Not Specified",
                "Link of the Job": direct_link,
                "Area": job.get("location") or job.get("city") or "Remote",
                "Link to Apply": direct_link,
                "Place You Got That Job From": job.get("source_platform") or job.get("source") or "Company Career Page",
                "Salary Quotation if Given": job.get("salary", "Not Disclosed"),
                "Company Name": job.get("company", ""),
                "Job ID": job.get("id", ""),
                "Work Mode": job.get("work_mode", "Onsite"),
                "Posted Date": job.get("posted_date", "Last 24h"),
                "Freshness Hours": job.get("freshness_hours", 24),
                "Experience Required": job.get("experience_required", "0-2 years"),
                "Tech Stack": job.get("tech_stack", ""),
                "Skills Required": job.get("skills_required", ""),
                "Employment Type": job.get("employment_type", "Full-time"),
                "Seniority Level": job.get("seniority_level", "Mid-Level"),
                "Responsibilities": job.get("responsibilities", ""),
                "Benefits": job.get("benefits", ""),
                "Recruiter Name": job.get("recruiter_name", ""),
                "Recruiter Email": job.get("recruiter_email", ""),
                "Hiring Manager Name": job.get("hiring_manager_name", ""),
                "Recruiter LinkedIn URL": job.get("recruiter_linkedin_url", ""),
                "Cold Email Subject": job.get("cold_email_subject", ""),
                "Custom Cold Email": job.get("custom_cold_email", ""),
                "Recruiter Contact Strategy": job.get("recruiter_contact_strategy", "")
            })
            
        df_all = pd.DataFrame(data_rows)
        if df_all.empty:
            # Create a mock dataframe with columns if empty
            df_all = pd.DataFrame(columns=cls.COLUMNS)
            
        # Segments by Target Cities / Remote
        city_sheets = {
            "Pune": df_all[df_all["City"].str.lower().str.contains("pune", na=False)],
            "Bangalore": df_all[df_all["City"].str.lower().str.contains("bangalore|bengaluru", na=False)],
            "Hyderabad": df_all[df_all["City"].str.lower().str.contains("hyderabad", na=False)],
            "Chennai": df_all[df_all["City"].str.lower().str.contains("chennai", na=False)],
            "Mumbai": df_all[df_all["City"].str.lower().str.contains("mumbai", na=False)],
            "Remote": df_all[df_all["City"].str.lower().str.contains("remote", na=False)],
        }
        
        # Openpyxl Excel writer
        with pd.ExcelWriter(filename, engine="openpyxl") as writer:
            for sheet_name, df_city in city_sheets.items():
                if df_city.empty:
                    # Write an empty structure with columns
                    df_city = pd.DataFrame(columns=cls.COLUMNS)
                else:
                    # Sort inside sheet: group by Onsite/Hybrid/Remote, newest first (Freshness Hours smallest)
                    df_city = df_city.sort_values(
                        by=["Work Mode", "Freshness Hours"], 
                        ascending=[True, True]
                    )
                
                # Write to sheet
                df_city.to_excel(writer, sheet_name=sheet_name, index=False)
                
                # Access Workbook sheet to style it
                workbook = writer.book
                worksheet = writer.sheets[sheet_name]
                
                # Enable Auto-Filter
                worksheet.auto_filter.ref = worksheet.dimensions
                
                # Freeze Header row
                worksheet.freeze_panes = "A2"
                
                # Define Styles
                header_font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
                header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid") # Dark Blue Accent
                regular_font = Font(name="Segoe UI", size=10)
                
                freshness_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid") # Soft Green
                salary_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid") # Soft Gold
                
                border_side = Side(border_style="thin", color="D9D9D9")
                cell_border = Border(left=border_side, right=border_side, top=border_side, bottom=border_side)
                
                # Style Header Row
                for col_idx in range(1, len(cls.COLUMNS) + 1):
                    cell = worksheet.cell(row=1, column=col_idx)
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=False)
                    
                # Style Regular Data Cells
                max_row = worksheet.max_row
                for r_idx in range(2, max_row + 1):
                    for c_idx in range(1, len(cls.COLUMNS) + 1):
                        cell = worksheet.cell(row=r_idx, column=c_idx)
                        cell.font = regular_font
                        cell.border = cell_border
                        
                        # Left align text descriptions, right align numeric/currency, center categories
                        header_name = cls.COLUMNS[c_idx - 1]
                        if header_name in ["Job Name", "Company Name", "JD", "Responsibilities", "Custom Cold Email", "Link of the Job", "Link to Apply"]:
                            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=False)
                        elif header_name in ["Job ID", "Work Mode", "Area", "Posted Date"]:
                            cell.alignment = Alignment(horizontal="center", vertical="center")
                        
                        # Apply highlights
                        if header_name == "Freshness Hours":
                            # Fresh jobs (< 24h) highlighted soft green
                            val = cell.value
                            try:
                                if val is not None and float(val) <= 24:
                                    cell.fill = freshness_fill
                            except Exception:
                                pass
                        elif header_name == "Salary Quotation if Given" and cell.value and cell.value != "Not Disclosed":
                            cell.fill = salary_fill
                            
                # Auto-fit Column widths to prevent truncation
                for col in worksheet.columns:
                    col_letter = get_column_letter(col[0].column)
                    # Limit width range between 12 and 45 to keep readable dimensions
                    max_len = 0
                    for cell in col:
                        if cell.value:
                            max_len = max(max_len, len(str(cell.value)))
                    worksheet.column_dimensions[col_letter].width = min(max(max_len + 3, 12), 45)
                    
        logger.info(f"Exporter: Workbook successfully generated and formatted: {filename}")
